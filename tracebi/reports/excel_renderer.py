"""
Excel renderer — produces well-formatted, multi-section .xlsx reports.

Features:
- Cover sheet with report metadata and parameter summary
- One worksheet per report (all sections on a single sheet) or per table
- Styled headers, alternating row colors, bold totals rows
- Embedded charts (bar, line, pie) as native Excel charts
- Column auto-sizing
- Frozen header rows

Requires: pip install openpyxl
"""

from __future__ import annotations

import logging
from datetime import datetime, timezone

import pandas as pd

from tracebi.reports.base_renderer import BaseRenderer
from tracebi.reports.report import (
    Report, SectionType,
    TextSection, TableSection, ChartSection,
    MetricSection, RowSection,
    resolve_number_format,
)

logger = logging.getLogger(__name__)


# ── Color palette ──────────────────────────────────────────────────────────
COLORS = {
    "header_fill":    "1F3864",   # dark navy
    "header_font":    "FFFFFF",
    "subheader_fill": "2F5496",
    "subheader_font": "FFFFFF",
    "alt_row":        "EBF0F7",
    "totals_fill":    "D6DCE4",
    "totals_font":    "1F3864",
    "border":         "B8CCE4",
    "cover_bg":       "1F3864",
    "cover_font":     "FFFFFF",
    "accent":         "2E74B5",
    "note_fill":      "FFF2CC",
    "callout_fill":   "DEEBF7",
}


class ExcelRenderer(BaseRenderer):
    """
    Renders a Report to a formatted .xlsx file.

    Usage:
        from tracebi.reports import ExcelRenderer
        renderer = ExcelRenderer()
        manifest = renderer.render(report, "output/sales_report.xlsx")
    """

    FORMAT = "excel"

    def __init__(self, include_cover: bool = True, include_lineage_sheet: bool = True):
        """
        Args:
            include_cover:         Add a styled cover sheet as the first tab.
            include_lineage_sheet: Add a 'Lineage' sheet listing all data sources.
        """
        self.include_cover = include_cover
        self.include_lineage_sheet = include_lineage_sheet
        self._width_overrides: dict[str, int] = {}

    def _render(self, report: Report, output_path: str) -> None:
        try:
            import openpyxl
            from openpyxl import Workbook
            from openpyxl.styles import (
                PatternFill, Font, Alignment, Border, Side,
                GradientFill,
            )
            from openpyxl.utils import get_column_letter
            from openpyxl.chart import BarChart, LineChart, PieChart, Reference
            from openpyxl.chart.series import DataPoint
        except ImportError:
            raise ImportError(
                "openpyxl is required for ExcelRenderer.\n"
                "Install with: pip install openpyxl"
            )

        wb = Workbook()
        wb.remove(wb.active)   # remove default sheet

        if self.include_cover:
            self._write_cover(wb, report)

        # Main report sheet
        ws = wb.create_sheet("Report")
        row = 1
        self._width_overrides = {}   # column letter → width, applied after autosize

        for section in report.sections:
            row = self._write_section(ws, section, row, wb)

        # Auto-size columns on the main sheet
        self._autosize_columns(ws)
        for letter, width in self._width_overrides.items():
            ws.column_dimensions[letter].width = width

        if self.include_lineage_sheet:
            self._write_lineage_sheet(wb, report)

        wb.save(output_path)

    def _write_section(self, ws, section, row: int, wb) -> int:
        """Dispatch a single section to its writer. RowSections have no
        side-by-side equivalent in Excel, so their children render stacked."""
        if section.section_type == SectionType.TEXT:
            return self._write_text(ws, section, row)
        elif section.section_type == SectionType.TABLE:
            return self._write_table(ws, section, row)
        elif section.section_type == SectionType.CHART:
            return self._write_chart(ws, section, row, wb)
        elif section.section_type == SectionType.SPACER:
            return row + section.height
        elif section.section_type == SectionType.METRICS:
            return self._write_metrics(ws, section, row)
        elif section.section_type == SectionType.ROW:
            for child in section.sections:
                row = self._write_section(ws, child, row, wb)
            return row
        return row

    # ── Metrics section ────────────────────────────────────────────────────

    def _write_metrics(self, ws, section: MetricSection, row: int) -> int:
        from openpyxl.styles import Font, PatternFill, Alignment

        if section.title:
            c = ws.cell(row=row, column=1, value=section.title)
            c.font = Font(bold=True, size=12,
                          color=COLORS["subheader_fill"], name="Calibri")
            row += 1

        label_row, value_row, delta_row = row, row + 1, row + 2
        ws.row_dimensions[value_row].height = 26
        has_delta = any(m.delta is not None for m in section.metrics)

        for col_idx, m in enumerate(section.metrics, 1):
            c = ws.cell(row=label_row, column=col_idx, value=m.label)
            c.font = Font(bold=True, size=9, color="5A6B8A", name="Calibri")
            c.alignment = Alignment(horizontal="left", vertical="bottom")

            c = ws.cell(row=value_row, column=col_idx, value=m.formatted_value())
            c.font = Font(bold=True, size=16,
                          color=COLORS["cover_bg"], name="Calibri")

            if m.delta is not None:
                up = m.delta >= 0
                good = up == m.good_when_up
                arrow = "▲" if up else "▼"
                c = ws.cell(row=delta_row, column=col_idx,
                            value=f"{arrow} {m.delta:+.1%}")
                c.font = Font(bold=True, size=9, name="Calibri",
                              color="2E7D32" if good else "C62828")

        return (delta_row if has_delta else value_row) + 2

    # ── Cover sheet ────────────────────────────────────────────────────────

    def _write_cover(self, wb, report: Report) -> None:
        from openpyxl.styles import PatternFill, Font, Alignment
        ws = wb.create_sheet("Cover", 0)
        ws.sheet_view.showGridLines = False

        def cell(row, col, value="", bold=False, size=11,
                 color="000000", bg=None, align="left", wrap=False):
            c = ws.cell(row=row, column=col, value=value)
            c.font = Font(bold=bold, size=size, color=color, name="Calibri")
            c.alignment = Alignment(horizontal=align, vertical="center",
                                    wrap_text=wrap)
            if bg:
                c.fill = PatternFill("solid", fgColor=bg)
            return c

        # Title block
        ws.row_dimensions[3].height = 40
        cell(3, 2, report.name, bold=True, size=24,
             color=COLORS["cover_font"], bg=COLORS["cover_bg"], align="center")
        ws.merge_cells("B3:H3")

        # Subtitle / description
        if report._description:
            ws.row_dimensions[4].height = 25
            cell(4, 2, report._description, size=12,
                 color="595959", align="center", wrap=True)
            ws.merge_cells("B4:H4")

        # Metadata table
        meta_start = 6
        items = [
            ("Author",     report._author or "—"),
            ("Generated",  datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")),
            ("Sections",   str(len(report.sections))),
        ]
        for k, v in report._parameters.items():
            items.append((k.replace("_", " ").title(), str(v)))

        for i, (label, value) in enumerate(items):
            r = meta_start + i
            ws.row_dimensions[r].height = 20
            cell(r, 2, label, bold=True, size=11,
                 color=COLORS["cover_font"], bg=COLORS["subheader_fill"])
            cell(r, 3, value, size=11)

        # Column widths
        ws.column_dimensions["A"].width = 3
        ws.column_dimensions["B"].width = 22
        ws.column_dimensions["C"].width = 40
        for col in "DEFGH":
            ws.column_dimensions[col].width = 15

    # ── Text section ───────────────────────────────────────────────────────

    def _write_text(self, ws, section: TextSection, row: int) -> int:
        from openpyxl.styles import Font, Alignment

        style = section.style
        if style == "heading1":
            ws.row_dimensions[row].height = 28
            c = ws.cell(row=row, column=1, value=section.title or section.content)
            c.font = Font(bold=True, size=16, color=COLORS["cover_bg"], name="Calibri")
            c.alignment = Alignment(vertical="center")
            row += 1
        elif style == "heading2":
            ws.row_dimensions[row].height = 22
            c = ws.cell(row=row, column=1, value=section.title or section.content)
            c.font = Font(bold=True, size=13, color=COLORS["subheader_fill"], name="Calibri")
            row += 1
        elif style in ("note", "callout"):
            bg = COLORS["note_fill"] if style == "note" else COLORS["callout_fill"]
            c = ws.cell(row=row, column=1, value=section.content)
            c.fill = __import__("openpyxl").styles.PatternFill("solid", fgColor=bg)
            c.font = Font(size=10, italic=True, name="Calibri")
            c.alignment = __import__("openpyxl").styles.Alignment(wrap_text=True)
            ws.merge_cells(
                start_row=row, start_column=1,
                end_row=row, end_column=8
            )
            row += 2
        else:
            if section.title:
                c = ws.cell(row=row, column=1, value=section.title)
                c.font = Font(bold=True, size=11, name="Calibri")
                row += 1
            c = ws.cell(row=row, column=1, value=section.content)
            c.font = Font(size=10, name="Calibri")
            c.alignment = __import__("openpyxl").styles.Alignment(wrap_text=True)
            row += 2

        return row

    # ── Table section ──────────────────────────────────────────────────────

    def _write_table(self, ws, section: TableSection, row: int) -> int:
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

        df = section.get_display_df()
        if df.empty:
            row += 1
            return row

        # Section title
        if section.title:
            c = ws.cell(row=row, column=1, value=section.title)
            c.font = Font(bold=True, size=12,
                          color=COLORS["subheader_fill"], name="Calibri")
            row += 1

        num_cols = len(df.columns)
        header_row = row

        thin = Side(style="thin", color=COLORS["border"])
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # Header
        for col_idx, col_name in enumerate(df.columns, 1):
            c = ws.cell(row=row, column=col_idx, value=str(col_name))
            c.font = Font(bold=True, size=10,
                          color=COLORS["header_font"], name="Calibri")
            c.fill = PatternFill("solid", fgColor=COLORS["header_fill"])
            c.alignment = Alignment(horizontal="center", vertical="center",
                                    wrap_text=True)
            c.border = border
        ws.row_dimensions[row].height = 18
        row += 1

        # Freeze header
        if section.freeze_header:
            ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

        # Compute totals before formatting
        totals: dict = {}
        if section.totals:
            for col in section.totals:
                display_col = (section.column_labels or {}).get(col, col)
                if display_col in df.columns:
                    try:
                        totals[display_col] = df[display_col].sum()
                    except Exception:
                        logger.warning(
                            "total for column %r could not be computed; "
                            "omitting from totals row", display_col, exc_info=True,
                        )

        # Data rows
        def _disp(col):
            return (section.column_labels or {}).get(col, col)

        fmt_map = {}
        if section.number_formats:
            for orig, disp in (section.column_labels or {}).items():
                if orig in section.number_formats:
                    fmt_map[disp] = section.number_formats[orig]
            fmt_map.update(section.number_formats)

        neg_cols = {_disp(c) for c in (section.highlight_negatives or [])}

        data_start_row = row
        for row_idx, (_, data_row) in enumerate(df.iterrows()):
            bg = COLORS["alt_row"] if row_idx % 2 == 1 else "FFFFFF"
            for col_idx, (col_name, val) in enumerate(data_row.items(), 1):
                c = ws.cell(row=row, column=col_idx)

                # Format numeric values
                display_val = val
                if col_name in fmt_map and pd.notna(val):
                    try:
                        fmt = fmt_map[col_name]
                        if isinstance(val, (int, float)):
                            display_val = val   # store raw; apply number_format
                            c.number_format = self._excel_number_format(fmt)
                    except Exception:
                        logger.warning(
                            "number_format %r failed for column %r; "
                            "writing unformatted value", fmt_map.get(col_name), col_name,
                        )

                c.value = display_val
                is_neg = (col_name in neg_cols and isinstance(val, (int, float))
                          and pd.notna(val) and val < 0)
                c.font = Font(size=9, name="Calibri",
                              color="C62828" if is_neg else "000000")
                c.fill = PatternFill("solid", fgColor=bg)
                c.alignment = Alignment(
                    horizontal="right" if isinstance(val, (int, float)) else "left",
                    vertical="center",
                )
                c.border = border
            ws.row_dimensions[row].height = 15
            row += 1
        data_end_row = row - 1

        # Color-scale (heat map) conditional formatting per column
        if section.color_scale and data_end_row >= data_start_row:
            from openpyxl.formatting.rule import ColorScaleRule
            from openpyxl.utils import get_column_letter
            col_positions = {c: i for i, c in enumerate(df.columns, 1)}
            for orig, hex_color in section.color_scale.items():
                disp = _disp(orig)
                if disp not in col_positions:
                    continue
                letter = get_column_letter(col_positions[disp])
                ws.conditional_formatting.add(
                    f"{letter}{data_start_row}:{letter}{data_end_row}",
                    ColorScaleRule(
                        start_type="min", start_color="FFFFFF",
                        end_type="max", end_color=hex_color.lstrip("#"),
                    ),
                )

        # Record explicit column widths (applied after autosize)
        if section.column_widths:
            from openpyxl.utils import get_column_letter
            col_positions = {c: i for i, c in enumerate(df.columns, 1)}
            for orig, width in section.column_widths.items():
                disp = _disp(orig)
                if disp in col_positions:
                    self._width_overrides[get_column_letter(col_positions[disp])] = width

        # Totals row
        if totals:
            for col_idx, col_name in enumerate(df.columns, 1):
                c = ws.cell(row=row, column=col_idx)
                val = totals.get(col_name, "")
                if col_idx == 1 and not val:
                    val = "Total"
                c.value = val
                c.font = Font(bold=True, size=9,
                              color=COLORS["totals_font"], name="Calibri")
                c.fill = PatternFill("solid", fgColor=COLORS["totals_fill"])
                c.border = border
                c.alignment = Alignment(horizontal="right" if isinstance(val, (int, float)) else "left")
            row += 1

        row += 1   # blank row after table
        return row

    # ── Chart section ──────────────────────────────────────────────────────

    def _write_chart(self, ws, section: ChartSection, row: int, wb) -> int:
        from openpyxl.chart import BarChart, LineChart, PieChart, Reference

        df = section.dataset.to_pandas() if section.dataset else pd.DataFrame()
        if df.empty or not section.x or not section.y:
            return row

        # Write a hidden data area for the chart to reference
        y_cols = [section.y] if isinstance(section.y, str) else list(section.y)
        data_cols = [section.x] + y_cols
        df_chart = df[data_cols].copy()

        data_start_row = row
        data_start_col = 12   # write chart data far right to stay out of the way

        # Headers
        for ci, col in enumerate(df_chart.columns, data_start_col):
            ws.cell(row=row, column=ci, value=col)
        row += 1

        # Data
        for _, dr in df_chart.iterrows():
            for ci, val in enumerate(dr, data_start_col):
                ws.cell(row=row, column=ci, value=val)
            row += 1

        data_end_row = row - 1
        n_cols = len(df_chart.columns)

        # Build chart
        chart_type = section.chart_type.lower()
        if chart_type in ("bar", "barh"):
            chart = BarChart()
            chart.type = "bar" if chart_type == "barh" else "col"
            chart.grouping = "clustered"
        elif chart_type == "line":
            chart = LineChart()
        elif chart_type == "pie":
            chart = PieChart()
        else:
            chart = BarChart()

        chart.title = section.title or ""
        if chart_type != "pie":  # PieChart has no x/y axes
            chart.y_axis.title = section.ylabel or (y_cols[0] if len(y_cols) == 1 else "")
            chart.x_axis.title = section.xlabel or section.x
        chart.width = 20
        chart.height = 12

        # Data reference (Y columns)
        data_ref = Reference(
            ws,
            min_col=data_start_col + 1,
            max_col=data_start_col + n_cols - 1,
            min_row=data_start_row,
            max_row=data_end_row,
        )
        chart.add_data(data_ref, titles_from_data=True)

        # Category reference (X column)
        cats = Reference(
            ws,
            min_col=data_start_col,
            min_row=data_start_row + 1,
            max_row=data_end_row,
        )
        chart.set_categories(cats)

        # Anchor chart below current report position
        anchor_row = data_start_row
        ws.add_chart(chart, f"A{anchor_row}")

        # Return row past the chart (approximate 25 rows for chart height)
        return anchor_row + 25

    # ── Lineage sheet ──────────────────────────────────────────────────────

    def _write_lineage_sheet(self, wb, report: Report) -> None:
        from openpyxl.styles import PatternFill, Font, Alignment

        ws = wb.create_sheet("Lineage")
        ws.sheet_view.showGridLines = False

        headers = ["Section", "Dataset", "Step", "Operation",
                   "Description", "Connector", "Source", "Timestamp"]
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=ci, value=h)
            c.font = Font(bold=True, size=9,
                          color=COLORS["header_font"], name="Calibri")
            c.fill = PatternFill("solid", fgColor=COLORS["header_fill"])
            c.alignment = Alignment(horizontal="center")

        row = 2
        for section in report.data_sections():
            if section.section_type not in (SectionType.TABLE, SectionType.CHART):
                continue
            ds = section.dataset
            if not ds:
                continue
            for step_idx, node in enumerate(ds.lineage, 1):
                conn = node.connector or {}
                ws.cell(row=row, column=1, value=section.title or "—")
                ws.cell(row=row, column=2, value=ds.name)
                ws.cell(row=row, column=3, value=step_idx)
                ws.cell(row=row, column=4, value=node.operation)
                ws.cell(row=row, column=5, value=node.description)
                ws.cell(row=row, column=6, value=conn.get("connector_name", ""))
                ws.cell(row=row, column=7, value=node.source or "")
                ws.cell(row=row, column=8, value=node.timestamp)
                for ci in range(1, 9):
                    ws.cell(row=row, column=ci).font = Font(size=9, name="Calibri")
                row += 1

        self._autosize_columns(ws)

    # ── Helpers ────────────────────────────────────────────────────────────

    @staticmethod
    def _autosize_columns(ws) -> None:
        from openpyxl.utils import get_column_letter
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    val = str(cell.value) if cell.value is not None else ""
                    max_len = max(max_len, len(val))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 8), 55)

    @staticmethod
    def _excel_number_format(fmt_string: str) -> str:
        """Convert a Python format string (or named shortcut) to an Excel
        number format string."""
        fmt_string = resolve_number_format(fmt_string)
        prefix = "$" if fmt_string.startswith("$") else ""
        if ".1%" in fmt_string:
            return '0.0%'
        elif "%" in fmt_string:
            return '0.00%'
        elif ",.2f" in fmt_string:
            return prefix + '#,##0.00'
        elif ",.0f" in fmt_string:
            return prefix + '#,##0'
        elif ".2f" in fmt_string:
            return prefix + '0.00'
        elif ".0f" in fmt_string:
            return prefix + '0'
        return 'General'
