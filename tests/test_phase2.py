"""
Tests for TraceBi Phase 2: Report Engine
"""

import os
import json
import tempfile
import pytest
import pandas as pd

from tracebi import DataModel, MemoryConnector
from tracebi.model.dataset import DataSet, LineageNode
from tracebi.reports.report import (
    Report, TextSection, TableSection, ChartSection,
    SpacerSection, SectionType, ReportManifest, RowSection,
)
from tracebi.reports.excel_renderer import ExcelRenderer
from tracebi.reports.html_renderer import HTMLRenderer


# ─────────────────────────────────────────────
# Fixtures
# ─────────────────────────────────────────────

def make_ds(name: str = "test") -> DataSet:
    df = pd.DataFrame({
        "region":  ["North", "South", "East"],
        "orders":  [100, 85, 120],
        "revenue": [5000.0, 4250.0, 6100.0],
    })
    node = LineageNode(
        operation="load", description=f"Load {name}",
        connector={"connector_name": "test", "connector_type": "CSV"},
        source=f"{name}.csv",
    )
    return DataSet(df=df, name=name, lineage=[node])


@pytest.fixture
def sample_report():
    ds = make_ds("sales")
    filtered_ds = ds.filter("orders > 90", description="High volume regions")
    return (
        Report("Test Report")
        .author("Test Author")
        .description("A test report.")
        .parameter("period", "Q1 2024")
        .add(TextSection(title="Summary", content="Summary heading", style="heading1"))
        .add(TextSection(content="Some body text.", style="normal"))
        .add(TextSection(content="A note here.", style="note"))
        .add(TableSection(
            title="Sales Table",
            dataset=filtered_ds,
            columns=["region", "orders", "revenue"],
            column_labels={"revenue": "Revenue ($)"},
            totals=["orders", "Revenue ($)"],
            number_formats={"revenue": "{:,.2f}"},
        ))
        .add(ChartSection(
            title="Revenue Chart",
            dataset=ds,
            chart_type="bar",
            x="region",
            y="revenue",
        ))
        .spacer()
    )


# ─────────────────────────────────────────────
# Report structure tests
# ─────────────────────────────────────────────

class TestReport:

    def test_fluent_builder(self, sample_report):
        assert sample_report.name == "Test Report"
        assert sample_report._author == "Test Author"
        assert sample_report._parameters == {"period": "Q1 2024"}
        assert len(sample_report.sections) == 6

    def test_section_types(self, sample_report):
        types = [s.section_type for s in sample_report.sections]
        assert SectionType.TEXT in types
        assert SectionType.TABLE in types
        assert SectionType.CHART in types
        assert SectionType.SPACER in types

    def test_build_manifest(self, sample_report):
        manifest = sample_report.build_manifest("excel", "/tmp/test.xlsx")
        assert manifest.report_name == "Test Report"
        assert manifest.format == "excel"
        assert manifest.output_path == "/tmp/test.xlsx"
        assert len(manifest.sections) == 6
        assert manifest.parameters == {"period": "Q1 2024"}

    def test_manifest_records_git_sha(self, sample_report):
        manifest = sample_report.build_manifest("excel", "/tmp/test.xlsx")
        # In a git checkout this is the HEAD SHA; outside one it's 'unknown'.
        assert manifest.git_sha
        assert manifest.git_sha == "unknown" or len(manifest.git_sha) == 40
        assert manifest.to_dict()["git_sha"] == manifest.git_sha

    def test_manifest_contains_lineage(self, sample_report):
        manifest = sample_report.build_manifest("html", "/tmp/test.html")
        # Find the table section manifest entry
        table_entry = next(
            (s for s in manifest.sections if s["section_type"] == "table"), None
        )
        assert table_entry is not None
        assert "dataset_lineage" in table_entry
        # The filtered_ds has 2 lineage nodes: load + filter
        assert len(table_entry["dataset_lineage"]) == 2

    def test_manifest_to_json(self, sample_report):
        manifest = sample_report.build_manifest("excel", "/tmp/test.xlsx")
        j = manifest.to_json()
        parsed = json.loads(j)
        assert parsed["report_name"] == "Test Report"
        assert isinstance(parsed["sections"], list)

    def test_table_section_get_display_df(self):
        ds = make_ds()
        section = TableSection(
            dataset=ds,
            columns=["region", "revenue"],
            column_labels={"revenue": "Rev"},
            max_rows=2,
        )
        df = section.get_display_df()
        assert list(df.columns) == ["region", "Rev"]
        assert len(df) == 2

    def test_describe_runs(self, sample_report, capsys):
        sample_report.describe()
        out = capsys.readouterr().out
        assert "Test Report" in out
        assert "Test Author" in out


# ─────────────────────────────────────────────
# Excel renderer tests
# ─────────────────────────────────────────────

class TestExcelRenderer:

    def test_renders_file(self, sample_report):
        with tempfile.TemporaryDirectory() as tmp:
            path = os.path.join(tmp, "report.xlsx")
            ExcelRenderer().render(report=sample_report, output_path=path)
            assert os.path.exists(path)
            assert os.path.getsize(path) > 1000

    def test_manifest_saved(self, sample_report):
        with tempfile.TemporaryDirectory() as tmp:
            path = os.path.join(tmp, "report.xlsx")
            ExcelRenderer().render(report=sample_report, output_path=path)
            manifest_path = path + ".manifest.json"
            assert os.path.exists(manifest_path)
            with open(manifest_path) as f:
                data = json.load(f)
            assert data["report_name"] == "Test Report"

    def test_manifest_returned(self, sample_report):
        with tempfile.TemporaryDirectory() as tmp:
            path = os.path.join(tmp, "report.xlsx")
            manifest = ExcelRenderer().render(report=sample_report, output_path=path)
            assert isinstance(manifest, ReportManifest)
            assert manifest.format == "excel"

    def test_renders_pie_chart(self):
        # Regression: PieChart has no x/y axes — setting axis titles crashed.
        report = Report("Pie Report").add(ChartSection(
            title="Revenue Share", dataset=make_ds("pie"),
            chart_type="pie", x="region", y="revenue",
        ))
        with tempfile.TemporaryDirectory() as tmp:
            path = os.path.join(tmp, "pie.xlsx")
            ExcelRenderer().render(report=report, output_path=path)
            assert os.path.getsize(path) > 1000

    def test_excel_has_sheets(self, sample_report):
        import openpyxl
        with tempfile.TemporaryDirectory() as tmp:
            path = os.path.join(tmp, "report.xlsx")
            ExcelRenderer(include_cover=True, include_lineage_sheet=True).render(
                report=sample_report, output_path=path
            )
            wb = openpyxl.load_workbook(path)
            assert "Cover" in wb.sheetnames
            assert "Report" in wb.sheetnames
            assert "Lineage" in wb.sheetnames

    def test_no_cover_option(self, sample_report):
        import openpyxl
        with tempfile.TemporaryDirectory() as tmp:
            path = os.path.join(tmp, "report.xlsx")
            ExcelRenderer(include_cover=False, include_lineage_sheet=False).render(
                report=sample_report, output_path=path
            )
            wb = openpyxl.load_workbook(path)
            assert "Cover" not in wb.sheetnames
            assert "Lineage" not in wb.sheetnames
            assert "Report" in wb.sheetnames

    def test_no_manifest_option(self, sample_report):
        with tempfile.TemporaryDirectory() as tmp:
            path = os.path.join(tmp, "report.xlsx")
            ExcelRenderer().render(report=sample_report, output_path=path,
                                   save_manifest=False)
            assert not os.path.exists(path + ".manifest.json")


# ─────────────────────────────────────────────
# HTML renderer tests
# ─────────────────────────────────────────────

class TestHTMLRenderer:

    def test_renders_file(self, sample_report):
        with tempfile.TemporaryDirectory() as tmp:
            path = os.path.join(tmp, "report.html")
            HTMLRenderer().render(report=sample_report, output_path=path)
            assert os.path.exists(path)
            assert os.path.getsize(path) > 2000

    def test_html_is_valid(self, sample_report):
        with tempfile.TemporaryDirectory() as tmp:
            path = os.path.join(tmp, "report.html")
            HTMLRenderer().render(report=sample_report, output_path=path)
            with open(path) as f:
                content = f.read()
            assert "<!DOCTYPE html>" in content
            assert "Test Report" in content
            assert "Test Author" in content
            assert "Q1 2024" in content

    def test_html_contains_table_data(self, sample_report):
        with tempfile.TemporaryDirectory() as tmp:
            path = os.path.join(tmp, "report.html")
            HTMLRenderer().render(report=sample_report, output_path=path)
            with open(path) as f:
                content = f.read()
            assert "Sales Table" in content
            assert "North" in content   # data from the dataset

    def test_html_contains_lineage(self, sample_report):
        with tempfile.TemporaryDirectory() as tmp:
            path = os.path.join(tmp, "report.html")
            HTMLRenderer().render(report=sample_report, output_path=path)
            with open(path) as f:
                content = f.read()
            assert "Data Lineage" in content
            assert "filter" in content

    def test_manifest_saved(self, sample_report):
        with tempfile.TemporaryDirectory() as tmp:
            path = os.path.join(tmp, "report.html")
            HTMLRenderer().render(report=sample_report, output_path=path)
            assert os.path.exists(path + ".manifest.json")

    def test_manifest_returned(self, sample_report):
        with tempfile.TemporaryDirectory() as tmp:
            path = os.path.join(tmp, "report.html")
            manifest = HTMLRenderer().render(report=sample_report, output_path=path)
            assert isinstance(manifest, ReportManifest)
            assert manifest.format == "html"

    def test_empty_report(self):
        report = Report("Empty")
        with tempfile.TemporaryDirectory() as tmp:
            path = os.path.join(tmp, "empty.html")
            HTMLRenderer().render(report=report, output_path=path)
            assert os.path.exists(path)

    def test_custom_manifest_path(self, sample_report):
        with tempfile.TemporaryDirectory() as tmp:
            path = os.path.join(tmp, "report.html")
            custom_mp = os.path.join(tmp, "custom_manifest.json")
            HTMLRenderer().render(
                report=sample_report, output_path=path,
                manifest_path=custom_mp
            )
            assert os.path.exists(custom_mp)


# ─────────────────────────────────────────────
# Formatting & layout tests
# ─────────────────────────────────────────────

class TestMetricSection:

    def make_metrics_report(self):
        from tracebi.reports.report import Metric, MetricSection
        return Report("Metrics Report").add(MetricSection(
            title="Key Metrics",
            metrics=[
                Metric("Total Revenue", 1250000, format="currency0", delta=0.12),
                Metric("Error Rate", 0.034, format="percent", delta=0.01,
                       good_when_up=False),
            ],
        ))

    def test_html_renders_cards(self):
        html = HTMLRenderer().to_html(self.make_metrics_report())
        assert "metric-card" in html
        assert "$1,250,000" in html
        assert "3.4%" in html
        # Positive delta on a good_when_up metric is good (green)…
        assert 'metric-delta good">▲ +12.0%' in html
        # …but on a good_when_up=False metric it is bad (red)
        assert 'metric-delta bad">▲ +1.0%' in html

    def test_excel_renders(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = os.path.join(tmp, "metrics.xlsx")
            ExcelRenderer().render(self.make_metrics_report(), path,
                                   save_manifest=False)
            from openpyxl import load_workbook
            ws = load_workbook(path)["Report"]
            values = [c.value for row in ws.iter_rows() for c in row if c.value]
            assert "Total Revenue" in values
            assert "$1,250,000" in values

    def test_manifest_records_metrics(self):
        manifest = self.make_metrics_report().build_manifest("html", "x.html")
        assert manifest.sections[0]["metrics"][0]["label"] == "Total Revenue"

    def test_fluent_shortcut(self):
        from tracebi.reports.report import Metric
        report = Report("R").metrics([Metric("Orders", 10)], title="KPIs")
        assert report.sections[0].section_type == SectionType.METRICS


class TestRowSection:

    def make_row_report(self):
        from tracebi.reports.report import RowSection
        ds = make_ds("sales")
        return Report("Row Report").add(RowSection(sections=[
            TableSection(title="Left Table", dataset=ds),
            ChartSection(title="Right Chart", dataset=ds,
                         chart_type="bar", x="region", y="revenue"),
        ]))

    def test_html_side_by_side(self):
        html = HTMLRenderer().to_html(self.make_row_report())
        assert "layout-row" in html
        assert "Left Table" in html
        assert "Right Chart" in html

    def test_nested_lineage_in_html(self):
        html = HTMLRenderer().to_html(self.make_row_report())
        assert "Data Lineage" in html
        assert "Load sales" in html

    def test_data_sections_flattens(self):
        report = self.make_row_report()
        types = [s.section_type for s in report.data_sections()]
        assert types == [SectionType.TABLE, SectionType.CHART]

    def test_excel_renders_stacked(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = os.path.join(tmp, "row.xlsx")
            ExcelRenderer().render(self.make_row_report(), path,
                                   save_manifest=False)
            from openpyxl import load_workbook
            wb = load_workbook(path)
            values = [c.value for row in wb["Report"].iter_rows()
                      for c in row if c.value]
            assert "Left Table" in values
            # Nested dataset lineage reaches the Lineage sheet
            lineage_vals = [c.value for row in wb["Lineage"].iter_rows()
                            for c in row if c.value]
            assert "Load sales" in lineage_vals

    def test_fluent_shortcut(self):
        ds = make_ds()
        report = Report("R").row(TableSection(dataset=ds), widths=[2, 1])
        assert report.sections[0].section_type == SectionType.ROW


class TestTableStyling:

    def make_styled_ds(self):
        df = pd.DataFrame({
            "region": ["North", "South", "East"],
            "margin": [120.0, -45.0, 300.0],
        })
        return DataSet(df=df, name="margins",
                       lineage=[LineageNode(operation="load", description="Load")])

    def test_named_number_format_html(self):
        report = Report("R").add(TableSection(
            dataset=make_ds(), number_formats={"revenue": "currency"}))
        html = HTMLRenderer().to_html(report)
        assert "$5,000.00" in html

    def test_highlight_negatives_html(self):
        report = Report("R").add(TableSection(
            dataset=self.make_styled_ds(), highlight_negatives=["margin"]))
        html = HTMLRenderer().to_html(report)
        assert 'class="num neg"' in html
        # Only the one negative value is highlighted
        assert html.count('class="num neg"') == 1

    def test_color_scale_html(self):
        report = Report("R").add(TableSection(
            dataset=self.make_styled_ds(), color_scale={"margin": "#2E74B5"}))
        html = HTMLRenderer().to_html(report)
        # Max value gets the full color, min gets white
        assert "background:rgb(46,116,181)" in html
        assert "background:rgb(255,255,255)" in html

    def test_column_widths_html(self):
        report = Report("R").add(TableSection(
            dataset=make_ds(), column_widths={"region": 20}))
        html = HTMLRenderer().to_html(report)
        assert "min-width:140px" in html

    def test_excel_styling(self):
        report = Report("R").add(TableSection(
            dataset=self.make_styled_ds(),
            highlight_negatives=["margin"],
            color_scale={"margin": "#2E74B5"},
            column_widths={"region": 30},
            number_formats={"margin": "currency"},
        ))
        with tempfile.TemporaryDirectory() as tmp:
            path = os.path.join(tmp, "styled.xlsx")
            ExcelRenderer().render(report, path, save_manifest=False)
            from openpyxl import load_workbook
            ws = load_workbook(path)["Report"]
            # column_widths override autosize
            assert ws.column_dimensions["A"].width == 30
            # negative value in red font
            cells = {c.value: c for row in ws.iter_rows() for c in row}
            assert "C62828" in str(cells[-45.0].font.color.rgb)
            assert cells[-45.0].number_format == "$#,##0.00"
            # conditional formatting registered for the margin column
            assert len(ws.conditional_formatting._cf_rules) == 1

    def test_excel_named_format_mapping(self):
        fmt = ExcelRenderer._excel_number_format
        assert fmt("currency") == "$#,##0.00"
        assert fmt("currency0") == "$#,##0"
        assert fmt("percent") == "0.0%"
        assert fmt("comma") == "#,##0"
        assert fmt("{:,.2f}") == "#,##0.00"


class TestChartEnhancements:

    def test_area_chart_html(self):
        report = Report("R").add(ChartSection(
            dataset=make_ds(), chart_type="area", x="region", y="revenue"))
        html = HTMLRenderer().to_html(report)
        assert '<svg class="tb-chart' in html   # inline SVG, not a base64 PNG

    def test_show_values_bar(self):
        report = Report("R").add(ChartSection(
            dataset=make_ds(), chart_type="bar", x="region", y="revenue",
            show_values=True))
        html = HTMLRenderer().to_html(report)
        assert '<svg class="tb-chart' in html   # inline SVG, not a base64 PNG


class TestReportNotebookIntegration:

    def test_report_repr_html_is_iframe(self, sample_report):
        html = sample_report._repr_html_()
        assert html.strip().startswith("<iframe srcdoc=")
        # Report content is embedded (entity-escaped)
        assert "Test Report" in html

    def test_report_help_prints(self, capsys):
        Report("R").help()
        out = capsys.readouterr().out
        assert ".table(" in out
        assert ".metrics(" in out
        assert ".row(" in out


class TestSectionValidationAtConstruction:
    """
    Renderers used to substitute a default for anything unrecognised, so a
    typo produced a plausible-looking wrong report and exit code 0.
    Validating in __post_init__ surfaces the error at the line that wrote it.
    """

    def test_unknown_chart_type_rejected(self):
        with pytest.raises(ValueError, match="chart_type"):
            ChartSection(chart_type="stacked_bar")

    def test_unknown_chart_type_suggests_a_close_match(self):
        with pytest.raises(ValueError, match="Did you mean 'bar'"):
            ChartSection(chart_type="barr")

    def test_unknown_text_style_rejected(self):
        with pytest.raises(ValueError, match="Did you mean 'heading1'"):
            TextSection(content="x", style="headline1")

    def test_unknown_table_style_rejected(self):
        with pytest.raises(ValueError, match="Did you mean 'striped'"):
            TableSection(style="stripey")

    def test_valid_values_still_accepted(self):
        for t in ("bar", "barh", "line", "area", "pie", "scatter"):
            assert ChartSection(chart_type=t).chart_type == t
        for s in ("normal", "heading1", "heading2", "note", "callout"):
            assert TextSection(style=s).style == s


class TestChartSectionNormalisation:
    """Normalised at construction so a section round-trips through JSON."""

    def test_scalar_y_becomes_a_list(self):
        assert ChartSection(chart_type="bar", y="revenue").y == ["revenue"]

    def test_list_y_is_preserved(self):
        assert ChartSection(chart_type="bar", y=["a", "b"]).y == ["a", "b"]

    def test_figsize_list_becomes_a_tuple(self):
        assert ChartSection(chart_type="bar", figsize=[8, 4]).figsize == (8, 4)


class TestHeadingContentNotDiscarded:
    """
    heading1/heading2 used to render `title or content`, silently dropping
    content when both were set. Many call sites worked around it by passing
    the same string twice, so content renders only when it actually differs.
    """

    @staticmethod
    def _html(section):
        from tracebi.reports.html_renderer import HTMLRenderer
        return HTMLRenderer().to_html(Report("T").add(section))

    def test_distinct_content_is_recovered(self):
        html = self._html(TextSection(title="Summary",
                                      content="Sales rose 12% QoQ.",
                                      style="heading1"))
        assert "Summary" in html
        assert "Sales rose 12% QoQ." in html

    def test_duplication_idiom_does_not_double_print(self):
        html = self._html(TextSection(title="Revenue by Region",
                                      content="Revenue by Region",
                                      style="heading2"))
        assert html.count("Revenue by Region") == 1

    def test_content_only_heading_unchanged(self):
        html = self._html(TextSection(content="Just A Heading", style="heading1"))
        assert html.count("Just A Heading") == 1


class TestNestedRowLineage:
    def test_data_sections_recurses_through_nested_rows(self):
        ds = DataSet(pd.DataFrame({"a": [1]}), name="d")
        inner = RowSection(sections=[TableSection(dataset=ds)])
        outer = RowSection(sections=[inner, TableSection(dataset=ds)])
        report = Report("N").add(outer)

        with_data = [s for s in report.data_sections()
                     if getattr(s, "dataset", None) is not None]
        assert len(with_data) == 2, (
            "a doubly-nested RowSection must not vanish from the lineage walk"
        )

    def test_flat_row_behaviour_unchanged(self):
        ds = DataSet(pd.DataFrame({"a": [1]}), name="d")
        report = Report("F").add(RowSection(sections=[TableSection(dataset=ds)]))
        assert len(report.data_sections()) == 1


class TestChartFailuresRaise:
    def test_missing_column_raises_instead_of_drawing_the_error(self):
        from tracebi.reports.html_renderer import HTMLRenderer

        ds = DataSet(pd.DataFrame({"region": ["NE"], "revenue": [1.0]}), name="d")
        bad = ChartSection(chart_type="bar", dataset=ds, x="nope", y="revenue")
        with pytest.raises(ValueError, match="not present in its dataset"):
            HTMLRenderer().to_html(Report("C").add(bad))

    def test_missing_column_error_suggests_a_close_match(self):
        from tracebi.reports.html_renderer import HTMLRenderer

        ds = DataSet(pd.DataFrame({"region": ["NE"], "revenue": [1.0]}), name="d")
        bad = ChartSection(chart_type="bar", dataset=ds, x="regoin", y="revenue")
        with pytest.raises(ValueError, match="did you mean 'region'"):
            HTMLRenderer().to_html(Report("C").add(bad))


class TestManifestReceiptCompleteness:
    def test_text_manifest_carries_a_content_hash(self):
        d = TextSection(title="H", content="body").to_manifest_dict()
        assert len(d["content_sha256"]) == 64
        other = TextSection(title="H", content="tampered").to_manifest_dict()
        assert d["content_sha256"] != other["content_sha256"]

    def test_chart_manifest_is_re_renderable(self):
        d = ChartSection(chart_type="line", x="m", y="rev",
                         palette=["#fff"], show_values=True).to_manifest_dict()
        for key in ("chart_type", "x", "y", "color", "xlabel", "ylabel",
                    "figsize", "style", "palette", "show_values"):
            assert key in d

    def test_section_id_round_trips_when_set(self):
        assert TextSection(content="x", id="exec-summary").to_manifest_dict()["id"] == "exec-summary"
        assert "id" not in TextSection(content="x").to_manifest_dict()


# ── Report specs (reports as data) ────────────────────────────────────────────

def _spec_model():
    """A model with declared measures, for spec tests."""
    orders = pd.DataFrame({
        "order_id": [1, 2, 3, 4], "customer_id": [1, 2, 1, 2],
        "revenue": [100.0, 200.0, 300.0, 400.0],
        "cost": [60.0, 150.0, 180.0, 320.0],
    })
    customers = pd.DataFrame({"customer_id": [1, 2], "region": ["West", "NE"]})
    m = DataModel("Sales").add_connector(
        MemoryConnector("mem", {"orders": orders, "customers": customers})
    )
    m.add_table("orders", connector="mem", source="orders")
    m.add_table("customers", connector="mem", source="customers")
    m.add_dimension("dim_customer", table_name="customers",
                    key_col="customer_id", attributes=["region"])
    m.add_fact("fact_orders", table_name="orders", measures=["revenue", "cost"],
               foreign_keys={"dim_customer": "customer_id"})
    m.add_measure("revenue", column="revenue", agg="sum")
    m.add_measure("gross_margin", expr="revenue - cost", agg="sum")
    m.connect()
    return m


VALID_SPEC = {
    "name": "Regional Margin",
    "author": "an agent",
    "sections": [
        {"type": "text", "title": "Summary", "style": "heading1"},
        {"type": "table", "title": "By Region",
         "data": {"model": "Sales", "query": {
             "fact": "fact_orders", "measures": ["revenue", "gross_margin"],
             "dimensions": ["dim_customer.region"]}}},
    ],
}


class TestReportSpecRoundTrip:
    def test_round_trips_through_json(self):
        from tracebi.spec import ReportSpec

        spec = ReportSpec.from_dict(VALID_SPEC)
        assert ReportSpec.from_json(spec.to_json()).to_dict() == spec.to_dict()

    def test_builds_a_live_report(self):
        from tracebi.spec import ReportSpec

        report = ReportSpec.from_dict(VALID_SPEC).build({"Sales": _spec_model()})
        assert report.name == "Regional Margin"
        assert len(report.sections) == 2

    def test_built_report_renders_with_real_data(self):
        from tracebi.reports.html_renderer import HTMLRenderer
        from tracebi.spec import ReportSpec

        report = ReportSpec.from_dict(VALID_SPEC).build({"Sales": _spec_model()})
        html = HTMLRenderer().to_html(report)
        assert "West" in html and "NE" in html

    def test_report_to_spec_recovers_the_data_reference(self):
        from tracebi.reports.report import Report, TableSection
        from tracebi.spec import ReportSpec

        model = _spec_model()
        ds = model.query(fact="fact_orders", measures=["revenue"],
                         dimensions=["dim_customer.region"])
        report = Report("RT").add(TableSection(title="Rows", dataset=ds))

        spec = ReportSpec.from_report(report)
        assert spec.data_coverage() == {
            "total": 1, "with_data_ref": 1, "presentation_only": [],
        }
        assert spec.to_dict()["sections"][0]["data"]["model"] == "Sales"

    def test_rebuilt_report_matches_the_original(self):
        from tracebi.reports.html_renderer import HTMLRenderer
        from tracebi.reports.report import Report, TableSection
        from tracebi.spec import ReportSpec

        model = _spec_model()
        ds = model.query(fact="fact_orders", measures=["revenue"],
                         dimensions=["dim_customer.region"])
        original = Report("RT").add(TableSection(title="Rows", dataset=ds))
        rebuilt = ReportSpec.from_report(original).build({"Sales": model})

        renderer = HTMLRenderer()
        assert len(renderer.to_html(original)) == len(renderer.to_html(rebuilt))

    def test_adhoc_data_is_reported_as_not_declarative(self):
        """
        A dataset built from arbitrary transforms has no declarative form.
        Saying so is better than pretending it round-trips.
        """
        from tracebi.reports.report import Report, TableSection
        from tracebi.spec import ReportSpec

        model = _spec_model()
        report = Report("A").add(
            TableSection(title="T", dataset=model.load("orders"))
        )
        coverage = ReportSpec.from_report(report).data_coverage()
        assert coverage["with_data_ref"] == 0
        assert coverage["presentation_only"] == ["T"]

    def test_nested_rows_round_trip(self):
        from tracebi.spec import ReportSpec

        spec = ReportSpec.from_dict({
            "name": "Nested",
            "sections": [{"type": "row", "sections": [
                {"type": "row", "sections": [{"type": "text", "content": "deep"}]},
                {"type": "text", "content": "shallow"},
            ]}],
        })
        report = spec.build({})
        assert "deep" in ReportSpec.from_report(report).to_json()

    def test_metrics_round_trip(self):
        from tracebi.spec import ReportSpec

        spec = ReportSpec.from_dict({
            "name": "KPIs",
            "sections": [{"type": "metrics", "title": "Q2", "metrics": [
                {"label": "Revenue", "value": 1000, "format": "currency"},
            ]}],
        })
        report = spec.build({})
        again = ReportSpec.from_report(report).to_dict()
        assert again["sections"][0]["metrics"][0]["label"] == "Revenue"


class TestReportSpecValidation:
    """
    The point of a spec is being checkable before it runs. An agent that
    gets a field-scoped error fixes its output; one that gets a rendered
    wrong report does not.
    """

    @staticmethod
    def _errors(spec_dict, models=None):
        from tracebi.spec import ReportSpec

        return ReportSpec.from_dict(spec_dict).validate(
            models if models is not None else {"Sales": _spec_model()}
        )["errors"]

    def test_valid_spec_passes(self):
        assert self._errors(VALID_SPEC) == []

    def test_bad_chart_type_is_caught_without_executing(self):
        errs = self._errors({"name": "x", "sections": [
            {"type": "chart", "chart_type": "stacked_bar"}]})
        assert any("chart_type" in e and "sections[0]" in e for e in errs)

    def test_bad_style_suggests_a_match(self):
        errs = self._errors({"name": "x", "sections": [
            {"type": "text", "style": "headline1"}]})
        assert any("heading1" in e for e in errs)

    def test_unknown_section_type_suggests_a_match(self):
        errs = self._errors({"name": "x", "sections": [{"type": "tabel"}]})
        assert any("table" in e for e in errs)

    def test_unknown_field_suggests_a_match(self):
        errs = self._errors({"name": "x", "sections": [
            {"type": "text", "contnet": "typo"}]})
        assert any("content" in e for e in errs)

    def test_unknown_model_is_caught(self):
        errs = self._errors({"name": "x", "sections": [
            {"type": "table", "data": {"model": "Nope", "query": {
                "fact": "fact_orders", "measures": ["revenue"]}}}]})
        assert any("data.model" in e for e in errs)

    def test_unknown_fact_is_caught(self):
        errs = self._errors({"name": "x", "sections": [
            {"type": "table", "data": {"model": "Sales", "query": {
                "fact": "fact_nope", "measures": ["revenue"]}}}]})
        assert any("data.query.fact" in e for e in errs)

    def test_undeclared_measure_is_caught(self):
        errs = self._errors({"name": "x", "sections": [
            {"type": "table", "data": {"model": "Sales", "query": {
                "fact": "fact_orders", "measures": ["revenu"]}}}]})
        assert any("data.query.measures" in e for e in errs)

    def test_unknown_dimension_is_caught(self):
        errs = self._errors({"name": "x", "sections": [
            {"type": "table", "data": {"model": "Sales", "query": {
                "fact": "fact_orders", "measures": ["revenue"],
                "dimensions": ["dim_nope.region"]}}}]})
        assert any("data.query.dimensions" in e for e in errs)

    def test_errors_inside_nested_rows_carry_the_path(self):
        errs = self._errors({"name": "x", "sections": [
            {"type": "row", "sections": [{"type": "chart", "chart_type": "bogus"}]}]})
        assert any("sections[0].sections[0]" in e for e in errs)

    def test_missing_data_reference_warns_but_does_not_error(self):
        from tracebi.spec import ReportSpec

        result = ReportSpec.from_dict(
            {"name": "x", "sections": [{"type": "table", "title": "T"}]}
        ).validate({"Sales": _spec_model()})
        assert result["ok"] is True
        assert any("no data reference" in w for w in result["warnings"])

    def test_build_refuses_an_invalid_spec(self):
        from tracebi.spec import ReportSpec

        spec = ReportSpec.from_dict({"name": "x", "sections": [
            {"type": "chart", "chart_type": "bogus"}]})
        with pytest.raises(ValueError, match="not valid"):
            spec.build({"Sales": _spec_model()})

    def test_unknown_top_level_field_rejected(self):
        from tracebi.spec import ReportSpec

        with pytest.raises(ValueError, match="Unknown report spec field"):
            ReportSpec.from_dict({"name": "x", "titel": "typo"})

    def test_name_is_required(self):
        from tracebi.spec import ReportSpec

        with pytest.raises(ValueError, match="needs a 'name'"):
            ReportSpec.from_dict({"sections": []})


class TestReportSpecSchema:
    def test_covers_every_section_type(self):
        """
        Anti-drift: add a SectionType and forget the spec mapping, and this
        fails rather than silently producing an unbuildable schema.
        """
        from tracebi.reports.report import SectionType
        from tracebi.spec import SECTION_CLASSES

        assert set(SECTION_CLASSES) == {t.value for t in SectionType}

    def test_schema_is_json_serializable(self):
        import json

        from tracebi.spec import json_schema

        assert json.loads(json.dumps(json_schema()))

    def test_schema_publishes_enum_values(self):
        from tracebi.reports.report import CHART_TYPES
        from tracebi.spec import json_schema

        variants = {v["properties"]["type"]["const"]: v
                    for v in json_schema()["$defs"]["section"]["oneOf"]}
        assert variants["chart"]["properties"]["chart_type"]["enum"] == list(CHART_TYPES)

    def test_schema_omits_live_object_fields(self):
        """A spec references data; it never inlines a DataSet."""
        from tracebi.spec import json_schema

        variants = {v["properties"]["type"]["const"]: v
                    for v in json_schema()["$defs"]["section"]["oneOf"]}
        assert "dataset" not in variants["table"]["properties"]
        assert "data" in variants["table"]["properties"]


class TestChartSpecSvg:
    """
    Charts were base64 matplotlib PNGs: unreadable in a diff, unstylable,
    fixed-size, and unavailable at all without an optional dependency.
    """

    @staticmethod
    def _ds():
        return DataSet(pd.DataFrame({
            "region":  ["North", "South", "East", "West"],
            "revenue": [5000.0, -1200.0, 6100.0, 300.0],
            "cost":    [3000.0, 900.0, 4000.0, 100.0],
        }), name="d")

    def _svg(self, kind, **kw):
        from tracebi.reports.chart import ChartSpec

        y = kw.pop("y", ["revenue", "cost"])
        x = kw.pop("x", "region")
        return ChartSpec.from_section(
            ChartSection(dataset=self._ds(), chart_type=kind, x=x, y=y, **kw)
        ).to_svg()

    @pytest.mark.parametrize("kind", ["bar", "barh", "line", "area"])
    def test_multi_series_types_produce_valid_xml(self, kind):
        import xml.etree.ElementTree as ET

        ET.fromstring(self._svg(kind))

    @pytest.mark.parametrize("kind", ["pie", "scatter"])
    def test_single_series_types_produce_valid_xml(self, kind):
        import xml.etree.ElementTree as ET

        x = "cost" if kind == "scatter" else "region"
        ET.fromstring(self._svg(kind, x=x, y=["revenue"]))

    def test_negative_values_stay_inside_the_plot(self):
        """A negative bar must hang below the zero line, not off-canvas."""
        import xml.etree.ElementTree as ET

        root = ET.fromstring(self._svg("bar"))
        ns = "{http://www.w3.org/2000/svg}"
        for rect in root.iter(f"{ns}rect"):
            y, h = float(rect.get("y")), float(rect.get("height"))
            assert y >= 0 and y + h <= 420

    def test_labels_are_escaped(self):
        """A label is untrusted text; unescaped it would break the document."""
        import xml.etree.ElementTree as ET

        from tracebi.reports.chart import ChartSpec

        ds = DataSet(pd.DataFrame({
            "region": ["<script>alert(1)</script>", "a & b", '"q"'],
            "v": [1.0, 2.0, 3.0],
        }), name="x")
        svg = ChartSpec.from_section(
            ChartSection(dataset=ds, chart_type="bar", x="region", y=["v"])
        ).to_svg()
        ET.fromstring(svg)
        assert "<script>" not in svg

    def test_spec_round_trips_through_json(self):
        import json

        from tracebi.reports.chart import ChartSpec

        spec = ChartSpec.from_section(
            ChartSection(dataset=self._ds(), chart_type="bar",
                         x="region", y=["revenue"])
        )
        again = ChartSpec.from_dict(json.loads(json.dumps(spec.to_dict())))
        assert again.to_svg() == spec.to_svg()

    def test_empty_data_renders_a_placeholder_not_a_crash(self):
        from tracebi.reports.chart import ChartSpec

        ds = DataSet(pd.DataFrame({"r": [], "v": []}), name="e")
        svg = ChartSpec.from_section(
            ChartSection(dataset=ds, chart_type="bar", x="r", y=["v"])
        ).to_svg()
        assert "no data" in svg

    def test_svg_is_responsive_not_fixed_size(self):
        svg = self._svg("bar")
        assert "viewBox" in svg
        assert "width=" not in svg.split(">", 1)[0]   # no hardcoded px width

    def test_elements_carry_classes_for_theming(self):
        """A stylesheet must be able to restyle a chart without code changes."""
        svg = self._svg("bar")
        for cls in ("tb-chart", "tb-grid", "tb-tick", "tb-bar"):
            assert cls in svg

    def test_axis_ticks_are_round_numbers(self):
        from tracebi.reports.chart import ChartSpec

        ticks = ChartSpec._ticks(0, 6100)
        assert all(t == round(t) for t in ticks)
        assert len(ticks) >= 2

    def test_charts_render_without_matplotlib(self):
        """The whole point: no optional dependency for HTML charts."""
        import subprocess
        import sys

        code = (
            "import sys\n"
            "class B:\n"
            "    def find_spec(self, n, p=None, t=None):\n"
            "        if n.split('.')[0] == 'matplotlib':\n"
            "            raise ImportError(n)\n"
            "        return None\n"
            "sys.meta_path.insert(0, B())\n"
            "import pandas as pd\n"
            "from tracebi import DataSet\n"
            "from tracebi.reports.report import Report, ChartSection\n"
            "from tracebi.reports.html_renderer import HTMLRenderer\n"
            "ds = DataSet(pd.DataFrame({'r': ['a','b'], 'v': [1.0, 2.0]}), name='d')\n"
            "h = HTMLRenderer().to_html(Report('C').add(\n"
            "    ChartSection(dataset=ds, chart_type='bar', x='r', y='v')))\n"
            "assert 'matplotlib required' not in h\n"
            "assert 'tb-chart' in h\n"
            "print('ok')\n"
        )
        out = subprocess.run([sys.executable, "-c", code],
                             capture_output=True, text=True)
        assert out.returncode == 0, out.stderr
        assert "ok" in out.stdout

    def test_no_base64_png_in_html_output(self):
        html = HTMLRenderer().to_html(Report("C").add(ChartSection(
            title="T", dataset=self._ds(), chart_type="bar",
            x="region", y=["revenue"])))
        assert "data:image/png;base64," not in html
        assert '<svg class="tb-chart' in html


class TestThemeAndTemplateLayer:
    """
    The renderer held its stylesheet as a module constant and built the page
    with f-strings, so restyling or restructuring output meant forking it.
    """

    @staticmethod
    def _report():
        ds = DataSet(pd.DataFrame({"r": ["a", "b"], "v": [1.0, 2.0]}), name="d")
        return (Report("Themed")
                .add(TextSection(title="Head", style="heading1"))
                .add(TableSection(dataset=ds)))

    def test_default_output_is_unchanged(self):
        """The refactor must not alter what existing reports look like."""
        html = HTMLRenderer().to_html(self._report())
        assert "Segoe UI" in html          # default stylesheet still applied
        assert html.startswith("<!DOCTYPE html>")
        assert "</html>" in html.strip()[-10:]

    def test_theme_overrides_are_appended_so_they_win(self):
        from tracebi.reports.theme import Theme

        theme = Theme.default().with_overrides("body { background: #111; }")
        html = HTMLRenderer(theme=theme).to_html(self._report())
        assert "#111" in html
        # appended after the defaults, so equal specificity resolves to ours
        assert html.index("Segoe UI") < html.index("#111")

    def test_theme_can_replace_the_stylesheet_entirely(self):
        from tracebi.reports.theme import Theme

        html = HTMLRenderer(
            theme=Theme.from_css("body{font-family:monospace}", name="mono")
        ).to_html(self._report())
        assert "monospace" in html
        assert "Segoe UI" not in html

    def test_theme_from_file(self, tmp_path):
        from tracebi.reports.theme import Theme

        css = tmp_path / "brand.css"
        css.write_text("body { color: rebeccapurple; }")
        theme = Theme.from_file(css)
        assert theme.name == "brand"
        assert "rebeccapurple" in HTMLRenderer(theme=theme).to_html(self._report())

    def test_missing_theme_file_raises(self, tmp_path):
        from tracebi.reports.theme import Theme

        with pytest.raises(FileNotFoundError):
            Theme.from_file(tmp_path / "nope.css")

    def test_section_renderer_can_be_replaced(self):
        """The extension point that previously required editing the renderer."""
        html = HTMLRenderer(
            section_renderers={SectionType.TEXT: lambda s: "<aside>CUSTOM</aside>"}
        ).to_html(Report("R").add(TextSection(content="original")))
        assert "<aside>CUSTOM</aside>" in html
        assert "original" not in html

    def test_section_renderer_accepts_a_plain_string_key(self):
        html = HTMLRenderer(
            section_renderers={"text": lambda s: "<b>via string key</b>"}
        ).to_html(Report("R").add(TextSection(content="x")))
        assert "via string key" in html

    def test_head_and_body_extra_are_injected(self):
        html = HTMLRenderer(
            head_extra='<link rel="icon" href="/f.ico">',
            body_extra="<!-- analytics -->",
        ).to_html(self._report())
        assert 'rel="icon"' in html
        assert "<!-- analytics -->" in html

    def test_custom_template_takes_over_the_page_shell(self):
        template = (
            "<!DOCTYPE html><html><head><title>{{ title }}</title>"
            "<style>{{ css }}</style></head><body class=\"corp\">"
            "<header>ACME — {{ report.name }}</header>{{ sections }}</body></html>"
        )
        html = HTMLRenderer(template=template).to_html(self._report())
        assert '<body class="corp">' in html
        assert "ACME — Themed" in html

    def test_template_typo_fails_loudly(self):
        """
        StrictUndefined: a mistyped placeholder must raise, not silently
        render an empty string into a report someone will rely on.
        """
        with pytest.raises(Exception, match="undefined"):
            HTMLRenderer(template="<h1>{{ titl }}</h1>").to_html(Report("R"))

    def test_default_shell_needs_no_template_engine(self):
        """Reports must work on a base install; jinja2 is only for custom shells."""
        import subprocess
        import sys

        code = (
            "import sys\n"
            "class B:\n"
            "    def find_spec(self, n, p=None, t=None):\n"
            "        if n.split('.')[0] == 'jinja2':\n"
            "            raise ImportError(n)\n"
            "        return None\n"
            "sys.meta_path.insert(0, B())\n"
            "from tracebi.reports.report import Report, TextSection\n"
            "from tracebi.reports.html_renderer import HTMLRenderer\n"
            "h = HTMLRenderer().to_html(Report('R').add(TextSection(content='x')))\n"
            "assert h.startswith('<!DOCTYPE html>')\n"
            "try:\n"
            "    HTMLRenderer(template='{{ title }}').to_html(Report('R'))\n"
            "    raise AssertionError('custom template should need jinja2')\n"
            "except ImportError as e:\n"
            "    assert 'Jinja2' in str(e)\n"
            "print('ok')\n"
        )
        out = subprocess.run([sys.executable, "-c", code],
                             capture_output=True, text=True)
        assert out.returncode == 0, out.stderr
        assert "ok" in out.stdout


# ── Totals × column_labels ────────────────────────────────────────────────────

class TestTotalsWithRenamedColumns:
    """
    Renaming a column for presentation used to silently delete its total.

    get_display_df() returns display names, but the totals loop matched against
    the original names in section.totals, so the moment column_labels renamed a
    column its total became an empty cell — no error, no warning. A missing
    total reads as "none was requested" rather than "we lost it", which is the
    worst way for a financial report to be wrong, and it is exactly the
    combination anything composing a presentable report will produce.
    """

    def _total_cells(self, tmp_path, **kwargs):
        import re
        from tracebi.reports.html_renderer import HTMLRenderer
        from tracebi.reports.report import Report, TableSection
        out = tmp_path / "r.html"
        HTMLRenderer().render(Report("t").add(TableSection(**kwargs)), str(out))
        row = re.search(
            r"<tr[^>]*>(?:(?!</tr>).)*?<strong>Total</strong>.*?</tr>",
            out.read_text(), re.S,
        )
        assert row, "no totals row rendered"
        return [
            re.sub(r"<[^>]+>", "", c).strip()
            for c in re.findall(r"<td[^>]*>(.*?)</td>", row.group(0), re.S)
        ]

    @pytest.fixture
    def ds(self):
        import pandas as pd
        from tracebi.model.dataset import DataSet
        return DataSet(pd.DataFrame({"region": ["N", "S"], "amount": [10.0, 32.0]}), name="t")

    def test_total_without_labels(self, tmp_path, ds):
        cells = self._total_cells(tmp_path, dataset=ds, totals=["amount"])
        assert cells[-1] == "42"

    def test_total_survives_column_labels(self, tmp_path, ds):
        cells = self._total_cells(tmp_path, dataset=ds, totals=["amount"],
                                  column_labels={"amount": "Amount"})
        assert cells[-1] == "42", "renaming a column must not drop its total"

    def test_total_survives_labels_and_formats_together(self, tmp_path, ds):
        cells = self._total_cells(tmp_path, dataset=ds, totals=["amount"],
                                  column_labels={"amount": "Amount"},
                                  number_formats={"amount": "currency0"})
        assert cells[-1] == "$42"

    def test_totals_may_be_given_as_the_display_name(self, tmp_path, ds):
        # Either spelling works, so a caller who thinks in labels is not
        # silently ignored.
        cells = self._total_cells(tmp_path, dataset=ds, totals=["Amount"],
                                  column_labels={"amount": "Amount"})
        assert cells[-1] == "42"


class TestCustomSectionTypes:
    """
    HTMLRenderer.section_renderers accepts "a type the framework doesn't know",
    which is how a project adds a block without forking. _render_section already
    tolerated a plain-string section_type; to_manifest_dict did not, so a custom
    section rendered correctly and then crashed on the way to the manifest — the
    audit trail being the one artefact that must never be what breaks.
    """

    def _custom(self):
        from dataclasses import dataclass
        from tracebi.reports.report import ReportSection

        @dataclass
        class Banner(ReportSection):
            def __post_init__(self):
                self.section_type = "banner"

        return Banner(title="Hello")

    def test_manifest_accepts_a_string_section_type(self):
        d = self._custom().to_manifest_dict()
        assert d["section_type"] == "banner"
        assert d["title"] == "Hello"

    def test_builtin_enum_section_type_still_serialises(self):
        from tracebi.reports.report import TextSection
        assert TextSection(title="t").to_manifest_dict()["section_type"] == "text"

    def test_report_manifest_builds_with_a_custom_section(self, tmp_path):
        from tracebi.reports.html_renderer import HTMLRenderer
        from tracebi.reports.report import Report
        report = Report("r").add(self._custom())
        out = tmp_path / "r.html"
        # Rendering builds the manifest; before the fix this raised
        # AttributeError: 'str' object has no attribute 'value'.
        HTMLRenderer(section_renderers={"banner": lambda s: "<p>banner</p>"}).render(
            report, str(out))
        assert "banner" in out.read_text()


# ── Derived presentation defaults ─────────────────────────────────────────────

class TestDerivedDefaults:
    """
    A table rendered straight from a star-schema query used to show
    DIM_BRANCH.REGION as a heading and 1705495.2200000002 as a figure. Both
    were always fixable via column_labels and number_formats — but only by an
    author remembering, on every table. A human authoring one report notices;
    something composing at volume with nobody reading the output does not. So
    the defaults are derived from what the query already knows, and anything
    set explicitly still wins.
    """

    def _render(self, section, derive=True):
        import tempfile, os
        from tracebi.reports.html_renderer import HTMLRenderer
        from tracebi.reports.report import Report
        p = os.path.join(tempfile.mkdtemp(), "r.html")
        HTMLRenderer(derive_defaults=derive).render(Report("t").add(section), p)
        return open(p).read()

    @pytest.fixture
    def ds(self):
        import pandas as pd
        from tracebi.model.dataset import DataSet
        return DataSet(pd.DataFrame({
            "dim_branch.region": ["Midwest", "West"],
            "market_value": [1705495.2200000002, 2314506.39],
            "orders": [12, 8],
        }), name="t")

    # ── humanise ──
    def test_humanise_strips_dimension_addressing(self):
        from tracebi.reports.derive import humanise
        assert humanise("dim_branch.region") == "Region"
        assert humanise("dim_client.segment") == "Segment"

    def test_humanise_makes_snake_case_readable(self):
        from tracebi.reports.derive import humanise
        assert humanise("market_value") == "Market value"
        assert humanise("units") == "Units"

    def test_humanise_leaves_an_already_plain_name_alone(self):
        from tracebi.reports.derive import humanise
        assert humanise("region") == "Region"

    # ── rendering ──
    def test_headers_are_readable_without_the_author_saying_so(self, ds):
        from tracebi.reports.report import TableSection
        html = self._render(TableSection(dataset=ds))
        assert "Market value" in html
        assert "dim_branch.region" not in html

    def test_float_noise_does_not_reach_the_page(self, ds):
        from tracebi.reports.report import TableSection
        html = self._render(TableSection(dataset=ds))
        assert "1705495.2200000002" not in html
        assert "1,705,495.22" in html

    def test_whole_numbers_get_separators_not_decimals(self, ds):
        from tracebi.reports.report import TableSection
        html = self._render(TableSection(dataset=ds))
        assert ">12<" in html.replace(" ", "")

    def test_explicit_labels_still_win(self, ds):
        from tracebi.reports.report import TableSection
        html = self._render(TableSection(
            dataset=ds, column_labels={"market_value": "AUM (USD)"}))
        assert "AUM (USD)" in html
        assert "Market value" not in html

    def test_explicit_formats_still_win(self, ds):
        from tracebi.reports.report import TableSection
        html = self._render(TableSection(
            dataset=ds, number_formats={"market_value": "currency0"}))
        assert "$1,705,495" in html

    def test_derivation_can_be_turned_off(self, ds):
        # The previous raw output verbatim, for anyone depending on it.
        from tracebi.reports.report import TableSection
        html = self._render(TableSection(dataset=ds), derive=False)
        assert "dim_branch.region" in html
        assert "1705495.2200000002" in html

    def test_percent_suffix_is_recognised(self):
        import pandas as pd
        from tracebi.model.dataset import DataSet
        from tracebi.reports.derive import derive_number_formats
        df = pd.DataFrame({"margin_pct": [0.42, 0.51]})
        fmts = derive_number_formats(df, DataSet(df, name="t"))
        assert fmts["margin_pct"] == "percent"

    def test_get_display_df_is_unchanged_when_no_labels_passed(self, ds):
        # The renderer passes derived labels in; called directly, the section
        # behaves exactly as before.
        from tracebi.reports.report import TableSection
        cols = list(TableSection(dataset=ds).get_display_df().columns)
        assert cols == ["dim_branch.region", "market_value", "orders"]
